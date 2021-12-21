#!/usr/bin/env python
# coding: utf-8

# In[18]:


import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.optimize import curve_fit

book = load_workbook(r'C:\Users\Liaisan\Desktop\data.xlsx')
sheet = book.active
# y = kx + b
def func(x, k, b):
    return k*x + b
x = []
y = []
xx = []
yy = []
YYerr = []
nameX = sheet[1][0].value
nameY = sheet[1][1].value
for row in range(2,sheet.max_row+1):
    xx.append(sheet[row][0].value)
    yy.append(sheet[row][1].value)
    YYerr.append(sheet[row][2].value)
x = np.array(xx)
y = np.array(yy)
Yerr = np.array(YYerr)

popt, pcov = curve_fit(func, x, y) 
plt.plot(x, func(x, *popt), 'violet')
plt.xlabel(nameX)
plt.ylabel(nameY)
plt.grid()

plt.scatter(x, y)
plt.errorbar(x, y, yerr=Yerr,fmt='o')

plt.show()

print('y  =  kx + b')
print('k = ', popt[0])
print('b = ', popt[1])



